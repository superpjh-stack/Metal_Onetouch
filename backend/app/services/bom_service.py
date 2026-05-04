"""BOM(자재소요량) 생성 서비스"""
import uuid
from decimal import Decimal
from typing import Optional

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bom import BomHeader, BomItem
from app.models.quotation import Quotation, QuotationItem
from app.schemas.bom import BomRead


class BomService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def generate_from_quotation(
        self, quotation_id: uuid.UUID, created_by: uuid.UUID
    ) -> BomRead:
        quotation = (await self.db.execute(
            select(Quotation).where(Quotation.id == quotation_id)
        )).scalar_one_or_none()
        if not quotation:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Quotation not found")
        if quotation.status != "accepted":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="BOM은 확정(accepted) 견적에서만 생성 가능합니다",
            )

        existing = (await self.db.execute(
            select(BomHeader).where(BomHeader.quotation_id == quotation_id)
        )).scalar_one_or_none()
        if existing:
            await self.db.refresh(existing, ["items"])
            return BomRead.model_validate(existing)

        items_rows = (await self.db.execute(
            select(QuotationItem)
            .where(QuotationItem.quotation_id == quotation_id)
            .order_by(QuotationItem.sort_order)
        )).scalars().all()

        bom_items, total_weight = self._build_bom_items(items_rows)

        bom = BomHeader(
            quotation_id=quotation_id,
            order_id=quotation.order_id,
            total_weight_kg=total_weight,
            created_by=created_by,
        )
        self.db.add(bom)
        await self.db.flush()

        for i, item_data in enumerate(bom_items):
            bom_item = BomItem(
                bom_id=bom.id,
                material_code=item_data["material_code"],
                specification=item_data["specification"],
                quantity=item_data["quantity"],
                unit=item_data["unit"],
                unit_weight_kg=item_data.get("unit_weight_kg"),
                total_weight_kg=item_data["total_weight_kg"],
                sort_order=i,
            )
            self.db.add(bom_item)

        await self.db.flush()
        await self.db.refresh(bom, ["items"])
        return BomRead.model_validate(bom)

    def _build_bom_items(
        self, items: list[QuotationItem]
    ) -> tuple[list[dict], Decimal]:
        """quotation_items → BOM 라인 변환, 동일 재질 집계"""
        material_map: dict[str, dict] = {}
        total_weight = Decimal("0")

        for item in items:
            if item.item_type == "material":
                # description 예: "SUS304 — 200×150×3.2mm (1.234kg)"
                desc = item.description or ""
                material_code = desc.split("—")[0].strip() if "—" in desc else "UNKNOWN"
                weight_kg = Decimal(str(item.quantity))  # quantity는 kg 단위
                total_weight += weight_kg

                if material_code in material_map:
                    material_map[material_code]["quantity"] += weight_kg
                    material_map[material_code]["total_weight_kg"] += weight_kg
                else:
                    spec = desc.split("—")[1].strip() if "—" in desc else desc
                    material_map[material_code] = {
                        "material_code": material_code,
                        "specification": spec,
                        "quantity": weight_kg,
                        "unit": "kg",
                        "unit_weight_kg": None,
                        "total_weight_kg": weight_kg,
                    }

        bom_items = list(material_map.values())
        # sort by material_code
        bom_items.sort(key=lambda x: x["material_code"])
        return bom_items, total_weight

    async def get_bom(self, quotation_id: uuid.UUID) -> Optional[BomRead]:
        bom = (await self.db.execute(
            select(BomHeader).where(BomHeader.quotation_id == quotation_id)
        )).scalar_one_or_none()
        if not bom:
            return None
        await self.db.refresh(bom, ["items"])
        return BomRead.model_validate(bom)

    async def export_xlsx(self, bom_id: uuid.UUID) -> bytes:
        bom = (await self.db.execute(
            select(BomHeader).where(BomHeader.id == bom_id)
        )).scalar_one_or_none()
        if not bom:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="BOM not found")
        await self.db.refresh(bom, ["items"])

        try:
            import io as _io
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "BOM"

            # 헤더 정보
            ws.append(["BOM ID", str(bom.id)])
            ws.append(["견적 ID", str(bom.quotation_id)])
            ws.append(["리비전", bom.revision])
            ws.append([])

            # 컬럼 헤더
            header = ["재질코드", "규격", "수량", "단위", "단위중량(kg)", "총중량(kg)"]
            ws.append(header)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(fill_type="solid", fgColor="D3D3D3")

            # 데이터 행
            for item in bom.items:
                ws.append([
                    item.material_code,
                    item.specification,
                    float(item.quantity),
                    item.unit,
                    float(item.unit_weight_kg) if item.unit_weight_kg else "",
                    float(item.total_weight_kg),
                ])

            # 합계 행
            ws.append(["합계", "", "", "", "", float(bom.total_weight_kg)])
            last_row = ws.max_row
            for cell in ws[last_row]:
                cell.font = Font(bold=True)

            buf = _io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_501_NOT_IMPLEMENTED,
                detail="openpyxl이 설치되어 있지 않습니다",
            )
